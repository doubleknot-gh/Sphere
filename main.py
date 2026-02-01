
from fastapi import Depends, FastAPI, HTTPException, status, UploadFile, File
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional, List
import csv
import codecs
import os
import io
import zipfile
from openpyxl import load_workbook

from jose import JWTError, jwt
from passlib.context import CryptContext
from datetime import datetime, timedelta

# 내부 모듈 임포트
import database as db
from database import get_db
import models
from models import Member, MemberStatus

# 데이터베이스 테이블 생성 (실제 운영 환경에서는 Alembic 같은 마이그레이션 도구 사용 권장)
models.Base.metadata.create_all(bind=db.engine)

# --- 보안 및 인증 설정 ---

# 비밀번호 해싱 설정
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")

# JWT 설정
SECRET_KEY = os.getenv("SECRET_KEY", "a-very-secret-key-for-local-development")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 30

# DB 초기화용 비밀키
INIT_DB_SECRET = os.getenv("INIT_DB_SECRET", "local-init-secret")

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token")

# --- 유틸리티 함수 ---

def verify_password(plain_password, hashed_password):
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password):
    return pwd_context.hash(password)

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.utcnow() + expires_delta
    else:
        expire = datetime.utcnow() + timedelta(minutes=15)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

# --- Pydantic 모델 (데이터 유효성 검사) ---

class Token(BaseModel):
    access_token: str
    token_type: str

class TokenData(BaseModel):
    student_id: Optional[str] = None

class MemberInfo(BaseModel):
    name: str
    student_id: str
    club: Optional[str] = None
    status: MemberStatus
    role: str

class MemberCreate(BaseModel):
    student_id: str
    name: str
    club: str

class PasswordUpdate(BaseModel):
    current_password: str
    new_password: str

# --- FastAPI 애플리케이션 생성 ---

app = FastAPI()

# --- 인증 관련 의존성 함수 ---

async def get_current_user(token: str = Depends(oauth2_scheme), db_session: Session = Depends(get_db)):
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        student_id: str = payload.get("sub")
        if student_id is None:
            raise credentials_exception
        token_data = TokenData(student_id=student_id)
    except JWTError:
        raise credentials_exception
    
    user = db_session.query(Member).filter(Member.student_id == token_data.student_id).first()
    if user is None:
        raise credentials_exception
    return user

# [보안] 관리자 권한 의존성 주입
def get_current_admin(current_user: Member = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="관리자 권한이 필요합니다."
        )
    return current_user

# --- API 엔드포인트 ---

@app.get("/")
def read_root():
    return {"message": "Digital Membership API is running. Please access the Streamlit frontend application to use the service."}

@app.post("/token", response_model=Token)
async def login_for_access_token(form_data: OAuth2PasswordRequestForm = Depends(), db_session: Session = Depends(get_db)):
    """
    학번(username)과 비밀번호로 로그인하여 JWT 토큰 발급
    """
    user = db_session.query(Member).filter(Member.student_id == form_data.username).first()
    if not user or not verify_password(form_data.password, user.password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect student_id or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": user.student_id}, expires_delta=access_token_expires
    )
    return {"access_token": access_token, "token_type": "bearer"}


@app.get("/members/me", response_model=MemberInfo)
async def read_users_me(current_user: Member = Depends(get_current_user)):
    """
    인증된 사용자의 '이름, 학번, 상태' 정보 조회
    """
    return current_user

@app.put("/members/me/password")
def update_password(password_data: PasswordUpdate, db_session: Session = Depends(get_db), current_user: Member = Depends(get_current_user)):
    """
    인증된 사용자가 자신의 비밀번호를 변경
    """
    # 현재 비밀번호 확인
    if not verify_password(password_data.current_password, current_user.password):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Incorrect current password")
    
    # 새 비밀번호로 업데이트
    current_user.password = get_password_hash(password_data.new_password)
    db_session.commit()
    
    return {"message": "Password updated successfully"}

# --- 관리자 전용 API ---

@app.get("/admin/members", response_model=List[MemberInfo])
def read_all_members(db_session: Session = Depends(get_db), admin: Member = Depends(get_current_admin)):
    """
    [관리자] 전체 회원 목록 조회
    """
    return db_session.query(Member).order_by(Member.club, Member.student_id).all()

@app.post("/admin/upload-csv")
async def upload_csv(file: UploadFile = File(...), db_session: Session = Depends(get_db), admin: Member = Depends(get_current_admin)):
    """
    [관리자] CSV 또는 Excel 파일로 회원 일괄 등록 (형식: 이름,학번,소속동아리)
    """
    filename = file.filename.lower()
    rows = []

    content = await file.read()

    if filename.endswith(".csv"):
        try:
            decoded = content.decode('utf-8-sig')
        except UnicodeDecodeError:
            try:
                decoded = content.decode('cp949')
            except UnicodeDecodeError:
                raise HTTPException(status_code=400, detail="CSV 파일 인코딩 오류: UTF-8 또는 CP949 형식이 아닙니다.")

        reader = csv.reader(decoded.splitlines())
        for row in reader:
            if len(row) >= 3:
                # 입력 순서: 이름(0), 학번(1), 소속동아리(2) -> DB 처리용: (학번, 이름, 소속동아리)
                sid = row[1].strip()
                name = row[0].strip()
                club = row[2].strip()
                if sid and name:
                    rows.append((sid, name, club))
    
    elif filename.endswith(".xlsx"):
        try:
            wb = load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                if row and len(row) >= 3:
                    # 데이터가 없는 경우 방지 및 문자열 변환
                    # 입력 순서: 이름(0), 학번(1), 소속동아리(2)
                    name_val = row[0]
                    sid_val = row[1]
                    club_val = row[2]
                    
                    def clean_cell(val):
                        if val is None: return ""
                        s = str(val).strip()
                        # 엑셀에서 숫자가 20241234.0 처럼 읽히는 경우 .0 제거
                        if s.endswith('.0') and s.replace('.', '', 1).isdigit():
                            return s[:-2]
                        return s

                    name = clean_cell(name_val)
                    sid = clean_cell(sid_val)
                    club = clean_cell(club_val)
                    
                    if sid and name:
                        rows.append((sid, name, club))
        except zipfile.BadZipFile:
            raise HTTPException(status_code=400, detail="엑셀 파일 형식이 올바르지 않습니다. (혹시 CSV 파일의 확장자만 .xlsx로 바꾸셨나요? 엑셀에서 '다른 이름으로 저장'을 통해 .xlsx로 저장해주세요.)")
        except Exception as e:
            print(f"Excel Error: {str(e)}") # 로그에 에러 출력
            raise HTTPException(status_code=400, detail=f"엑셀 처리 중 오류가 발생했습니다: {str(e)}")
    else:
        raise HTTPException(status_code=400, detail="지원하지 않는 파일 형식입니다. .csv 또는 .xlsx 파일을 사용해주세요.")
    
    success_count = 0
    updated_count = 0
    skip_count = 0
    
    # [중요] 이번 요청에서 새로 추가된 회원을 임시 저장 (파일 내 중복 방지)
    new_members_cache = {}

    # [최적화] 파일에 있는 학번들을 미리 DB에서 조회하여 맵핑 (N+1 쿼리 방지 및 정확성 향상)
    # 헤더가 아닌 유효한 학번만 추출
    valid_sids = [r[0] for r in rows if not (any(h in r[0] for h in ["학번", "Student", "ID", "id"]) or any(h in r[1] for h in ["이름", "Name", "성명"]))]
    
    # DB에 존재하는 회원들을 한 번에 가져옴
    existing_members_query = db_session.query(Member).filter(Member.student_id.in_(valid_sids)).all()
    existing_members_map = {m.student_id: m for m in existing_members_query}

    try:
        for sid, name, club in rows:
            # 헤더 행 스킵 (학번이나 이름 자리에 제목이 들어있는 경우 건너뜀)
            # 예: '학번', 'Student ID', '이름', 'Name', '성명' 등이 포함되면 헤더로 간주
            if any(header in sid for header in ["학번", "Student", "ID", "id"]) or any(header in name for header in ["이름", "Name", "성명"]): continue

            # 1. DB에 이미 존재하는지 확인 (미리 가져온 맵에서 확인)
            if sid in existing_members_map:
                existing_member = existing_members_map[sid]
                # 이미 존재하는 경우: 소속 동아리 추가 (중복되지 않게)
                current_clubs = [c.strip() for c in (existing_member.club or "").split(',')]
                if club and club not in current_clubs:
                    if existing_member.club:
                        existing_member.club += f", {club}"
                    else:
                        existing_member.club = club
                    
                    # 총동아리연합회가 추가되면 관리자 권한 부여
                    if club == "총동아리연합회":
                        existing_member.role = "admin"
                    
                    updated_count += 1
                else:
                    skip_count += 1 # 이미 해당 동아리가 있거나 변경사항 없음
            
            # 2. DB에는 없지만, 이번 파일 업로드로 방금 추가된 회원인지 확인 (캐시 확인)
            elif sid in new_members_cache:
                member_to_update = new_members_cache[sid]
                current_clubs = [c.strip() for c in (member_to_update.club or "").split(',')]
                if club and club not in current_clubs:
                    if member_to_update.club:
                        member_to_update.club += f", {club}"
                    else:
                        member_to_update.club = club
                    
                    if club == "총동아리연합회" or name == "김근호":
                        member_to_update.role = "admin"
                    # 이미 success_count에 포함되었으므로 카운트는 유지하거나 updated로 이동 가능하나, 
                    # 사용자 혼동 방지를 위해 여기서는 별도 카운트 증가 없이 진행
            
            else:
                # 3. 완전히 새로운 회원 등록
                # 이름이 '김근호'이거나 소속 동아리가 '총동아리연합회'이면 관리자 권한 부여
                role = "admin" if club == "총동아리연합회" or name == "김근호" else "member"
                # 초기 비밀번호는 '1234'로 설정
                new_member = Member(
                    student_id=sid,
                    name=name,
                    club=club,
                    password=get_password_hash("1234"),
                    role=role,
                    status=MemberStatus.active
                )
                db_session.add(new_member)
                new_members_cache[sid] = new_member # 캐시에 등록
                success_count += 1
        
        db_session.commit()
        return {"message": f"신규 {success_count}명 등록, {updated_count}명 정보 업데이트 (변경없음 {skip_count}명)"}
    except Exception as e:
        db_session.rollback()
        print(f"Upload Error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"회원 등록 중 서버 오류가 발생했습니다: {str(e)}")

@app.post("/admin/members", response_model=MemberInfo)
def create_member(member: MemberCreate, db_session: Session = Depends(get_db), admin: Member = Depends(get_current_admin)):
    """
    [관리자] 개별 회원 직접 등록
    """
    if db_session.query(Member).filter(Member.student_id == member.student_id).first():
        raise HTTPException(status_code=400, detail="이미 존재하는 학번입니다.")
    
    # 이름이 '김근호'이거나 소속 동아리가 '총동아리연합회'이면 관리자 권한 부여
    role = "admin" if member.club == "총동아리연합회" or member.name == "김근호" else "member"
    
    new_member = Member(
        student_id=member.student_id,
        name=member.name,
        club=member.club,
        password=get_password_hash("1234"), # 초기 비밀번호는 1234로 설정
        role=role,
        status=MemberStatus.active
    )
    db_session.add(new_member)
    db_session.commit()
    db_session.refresh(new_member)
    return new_member

@app.patch("/admin/members/{student_id}/status")
def update_member_status(student_id: str, status: str, db_session: Session = Depends(get_db), admin: Member = Depends(get_current_admin)):
    user = db_session.query(Member).filter(Member.student_id == student_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    user.status = MemberStatus.active if status == "active" else MemberStatus.inactive
    db_session.commit()
    return {"message": "상태 변경 완료"}

@app.patch("/admin/members/{student_id}/club")
def update_member_club(student_id: str, club: str, db_session: Session = Depends(get_db), admin: Member = Depends(get_current_admin)):
    user = db_session.query(Member).filter(Member.student_id == student_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    user.club = club
    db_session.commit()
    return {"message": "소속 동아리 변경 완료"}

@app.patch("/admin/members/{student_id}/role")
def update_member_role(student_id: str, role: str, db_session: Session = Depends(get_db), admin: Member = Depends(get_current_admin)):
    user = db_session.query(Member).filter(Member.student_id == student_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    user.role = role
    db_session.commit()
    return {"message": "회원 권한 변경 완료"}

@app.delete("/admin/members/{student_id}")
def delete_member(student_id: str, db_session: Session = Depends(get_db), admin: Member = Depends(get_current_admin)):
    user = db_session.query(Member).filter(Member.student_id == student_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    db_session.delete(user)
    db_session.commit()
    return {"message": "삭제 완료"}

@app.patch("/admin/members/{student_id}/reset-password")
def reset_password(student_id: str, db_session: Session = Depends(get_db), admin: Member = Depends(get_current_admin)):
    """
    [관리자] 특정 회원의 비밀번호를 '1234'로 초기화
    """
    user = db_session.query(Member).filter(Member.student_id == student_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    user.password = get_password_hash("1234")
    db_session.commit()
    return {"message": f"Password for {student_id} has been reset to '1234'."}

@app.get("/init-db")
def init_database(secret: str, db_session: Session = Depends(get_db)):
    """
    [초기화] 데이터베이스에 관리자 및 테스트 계정 생성
    (Shell 접속이 어려울 때 브라우저에서 실행용)
    """
    messages = []

    # 비밀키가 일치하지 않으면 초기화 거부
    if secret != INIT_DB_SECRET:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Invalid secret for DB initialization")
    
    # 1. 관리자 계정 생성
    if not db_session.query(Member).filter(Member.student_id == "admin").first():
        admin_user = Member(
            student_id="admin",
            password=get_password_hash("admin1234"),
            name="김근호",
            club="총동아리연합회",
            status=MemberStatus.active,
            role="admin"
        )
        db_session.add(admin_user)
        messages.append("관리자 계정(admin) 생성 완료")
    
    # 2. 테스트 계정 생성
    if not db_session.query(Member).filter(Member.student_id == "20240001").first():
        test_user = Member(
            student_id="20240001",
            password=get_password_hash("1234"),
            name="김테스트",
            club="테니스부",
            status=MemberStatus.active,
            role="member"
        )
        db_session.add(test_user)
        messages.append("테스트 계정(20240001) 생성 완료")
        
    if not messages:
        messages.append("이미 모든 계정이 존재합니다.")
        
    db_session.commit()
    return {"status": "success", "details": messages}
